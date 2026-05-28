#!/usr/bin/env python3
"""
Alpha Schools — SY26/27 Portfolio Consolidation Builder & Refresher
====================================================================
Run this script any time you want to:
  1. Update the consolidation after changing a school's budget file
  2. Add a new school to the portfolio

Usage:
    python3 refresh_consolidation.py

Output:
    mnt/27 Budget/Alpha_Consolidation_SY2627.xlsx

Architecture:
    _Data tab  — raw P&L values for every school × scenario × row (hidden)
    Config tab — school list with Y/N include toggles (live filter)
    BASE tab   — portfolio P&L with SUMPRODUCT formulas → updates instantly
    CONSTRAINED tab — same
    STRETCH tab     — same
    Summary tab — side-by-side Y1 snapshot of all 19 schools (static)

    • Changing Y/N in Config filters the scenario tabs WITHOUT re-running this script
    • Re-run this script when school data changes or a new school is added
"""

import os, sys, shutil, subprocess
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
BUDGET_DIR = os.path.join(BASE_DIR, "mnt", "27 Budget")
OUT_FILE   = os.path.join(BUDGET_DIR, "Alpha_Consolidation_SY2627.xlsx")

# ── Palette ───────────────────────────────────────────────────────────────────
TITLE_BG  = "001F3864"
ADDR_BG   = "00E8E8E8"
HDR_BG    = "002F5496"
ENROLL_BG = "00D9E1F2"
YELLOW_BG = "00FFD966"
GREEN_BG  = "0070AD47"
DARK_BG   = "004F4F4F"
WHITE     = "00FFFFFF"
BLACK     = "00000000"
GRAY_TEXT = "00555555"
FMT_NUM   = '#,##0'
FMT_TEXT  = '@'

# ── School master list ────────────────────────────────────────────────────────
# To add a new school: append an entry here and re-run this script.
SCHOOLS = [
    {"name": "Alpha Chantilly",          "file": "Alpha_Chantilly_SY2627_Budget.xlsx",       "location": "Chantilly, VA",           "assumptions_url": "https://drive.google.com/file/d/1kFxXdDp30UJYX9ffpTaiNlFBQC6Osmxo/view"},
    {"name": "Alpha Charlotte",          "file": "Alpha_Charlotte_SY2627_Budget.xlsx",        "location": "Charlotte, NC",           "assumptions_url": "https://drive.google.com/file/d/1DNw60ave4pyHt9L_iaX_L0qIvHJOe9kK/view"},
    {"name": "Alpha Tampa",              "file": "Alpha_Tampa_SY2627_Budget.xlsx",            "location": "Tampa, FL",               "assumptions_url": "https://drive.google.com/file/d/1tSzrm0TVAagvUgvQf48bABEfMlefcHJJ/view"},
    {"name": "Alpha Palo Alto",          "file": "Alpha_PaloAlto_SY2627_Budget.xlsx",         "location": "Palo Alto, CA",           "assumptions_url": "https://drive.google.com/file/d/1ITbO_hqbIYOJIytHnrXP8NRbvCB56GB6/view"},
    {"name": "Alpha Boston",             "file": "Alpha_Boston_SY2627_Budget.xlsx",           "location": "Boston, MA",              "assumptions_url": "https://drive.google.com/file/d/1_754Tqwk-p9u1GFJyIllKw9aipWv7Def/view"},
    {"name": "Alpha Bethesda",           "file": "Alpha_Bethesda_SY2627_Budget.xlsx",         "location": "Bethesda, MD",            "assumptions_url": "https://drive.google.com/file/d/15iwRWzSc2huMVlLtd8owXMJgyfSkTj-2/view"},
    {"name": "Alpha New York",           "file": "Alpha_NewYork_SY2627_Budget.xlsx",          "location": "New York, NY",            "assumptions_url": "https://drive.google.com/file/d/1UXtPsSqjPEYeNtv3BOxxRsBXfnwGS20j/view"},
    {"name": "Alpha The Woodlands",      "file": "Alpha_Woodlands_SY2627_Budget.xlsx",        "location": "The Woodlands, TX",       "assumptions_url": "https://drive.google.com/file/d/1IHHwb5idV5uhqo5C3my2y2NXFTUIDILp/view"},
    {"name": "Alpha Keller",             "file": "Alpha_Keller_SY2627_Budget.xlsx",           "location": "Keller, TX",              "assumptions_url": "https://drive.google.com/file/d/1Oihy3QumH3ujJkd8nfH6w7Joflvy5QlM/view"},
    {"name": "Alpha Santa Monica",       "file": "Alpha_SantaMonica_SY2627_Budget.xlsx",      "location": "Santa Monica, CA",        "assumptions_url": "https://drive.google.com/file/d/16GCq21FPo-5YMSIn5aJ8C6oVlnvx4OGT/view"},
    {"name": "Alpha Dorado",             "file": "Alpha_Dorado_SY2627_Budget.xlsx",           "location": "Dorado, PR",              "assumptions_url": "https://drive.google.com/file/d/1YLwaW8qY6WZC6_uM5DvUmNHCsqsu8XHx/view"},
    {"name": "Alpha Raleigh",            "file": "Alpha_Raleigh_SY2627_Budget.xlsx",          "location": "Raleigh, NC",             "assumptions_url": "https://drive.google.com/file/d/1ZFYeY7rVs6OEwo7dZhiRiRf6hQJe48O5/view"},
    {"name": "Alpha Atlanta (Roswell)",  "file": "Alpha_Atlanta_Roswell_SY2627_Budget.xlsx",  "location": "Roswell, GA",             "assumptions_url": "https://drive.google.com/file/d/1jwsUj1o3ZmjXTehqyohDlHSBZXHE2tR5/view"},
    {"name": "Alpha Santa Barbara",      "file": "Alpha_SantaBarbara_SY2627_Budget.xlsx",     "location": "Santa Barbara, CA",       "assumptions_url": "https://drive.google.com/file/d/1SoFZ1fK8zQ1_TzaUswMYdeX1U_WxF9HU/view"},
    {"name": "Alpha Chicago",            "file": "Alpha_Chicago_SY2627_Budget.xlsx",          "location": "Chicago, IL",             "assumptions_url": "https://drive.google.com/file/d/169K3oxBpb7YjgENtTDe5pQemXy68heh8/view"},
    {"name": "Alpha La Jolla",           "file": "Alpha_LaJolla_SY2627_Budget.xlsx",          "location": "La Jolla, CA",            "assumptions_url": "https://drive.google.com/file/d/1dyQ9XWNPq1qDuYw3TFv4lNbTXQaUYRgi/view"},
    {"name": "Alpha Palm Beach Gardens", "file": "Alpha_PalmBeachGardens_SY2627_Budget.xlsx", "location": "Palm Beach Gardens, FL",  "assumptions_url": "https://drive.google.com/file/d/1Dwfb_5JRA2PjTbpf0hyTRFvRakoC3XJF/view"},
    {"name": "Alpha Kirkland",           "file": "Kirkland_SY2627_Budget_Model.xlsx",         "location": "Kirkland, WA",            "assumptions_url": "https://drive.google.com/file/d/1pe0ggBMCs1ReP7el2qZ_toJU_35_v8ML/view"},
    {"name": "Alpha Plano",              "file": "Plano_TX_SY2627_Budget_Model.xlsx",         "location": "Plano, TX",               "assumptions_url": "https://drive.google.com/file/d/17olBYQnvvrDFCdE7i5aQQ1dbnKw7sXfR/view"},
    {"name": "Alpha Houston",            "file": "Alpha_Houston_SY2627_Budget.xlsx",           "location": "Houston, TX",             "assumptions_url": "https://drive.google.com/file/d/1CVge2TkHplQC1arY_9Lcc-RZtS9xtAt-/view"},
    {"name": "Alpha Austin",             "file": "Alpha_Austin_SY2627_Budget.xlsx",            "location": "Austin, TX",              "assumptions_url": "https://drive.google.com/file/d/1OKNMdkd_Gb8US1kRQXMSp3AgtCwqMkXU/view"},
    {"name": "Alpha Miami",              "file": "Alpha_Miami_SY2627_Budget.xlsx",             "location": "Miami, FL",               "assumptions_url": "https://drive.google.com/file/d/1YSDnmSc_LMdUDq_Rg8IEzKXh55l2riOP/view"},
    {"name": "Alpha Fort Worth",         "file": None,                                         "location": "Fort Worth, TX",          "assumptions_url": "https://drive.google.com/file/d/1aV1Oatcl-AaKjFidnQRL2oxwudYOppzp/view",
     "comment": "needs target date and school model"},
]

N_SCHOOLS = len(SCHOOLS)
CONFIG_DATA_START = 4            # first school row in Config tab
CONFIG_DATA_END   = 70           # last row in Config range (room for up to 65 schools)

# ── P&L row definitions ───────────────────────────────────────────────────────
# (key, search_patterns_in_col_A, display_label, section, is_total)
# Patterns use startswith matching (case-insensitive) so they must be SPECIFIC
# enough not to match other rows (e.g., "EBIT  (" won't match "EBITDA").
PL_ROWS = [
    ("enrollment",  ["Enrolled Students"],                                       "Enrolled Students",               "enrollment", False),
    ("core_rev",    ["Core Tuition Revenue"],                                    "    Core Tuition Revenue",         "revenue",    False),
    ("ah_rev",      ["After-Hours Revenue"],                                     "    After-Hours Revenue",          "revenue",    False),
    ("ss_rev",      ["Summer School Revenue"],                                   "    Summer School Revenue",        "revenue",    False),
    ("total_rev",   ["TOTAL REVENUE"],                                           "TOTAL REVENUE",                    "revenue",    True ),
    ("lead_guides", ["Lead Guides"],                                             "    Lead Guides",                  "headcount",  False),
    ("hos",         ["Head of School"],                                          "    Head of School",               "headcount",  False),
    ("guides",      ["Guides"],                                                  "    Guides",                       "headcount",  False),
    ("admin",       ["Campus Administrator"],                                    "    Campus Administrator",         "headcount",  False),
    ("total_hc",    ["TOTAL HEADCOUNT"],                                         "TOTAL HEADCOUNT",                  "headcount",  True ),
    ("timeback",    ["Timeback"],                                                "    Timeback",                     "opex",       False),
    ("programs",    ["Programs", "Core Programs"],                               "    Programs",                     "opex",       False),
    ("misc",        ["Miscellaneous"],                                           "    Miscellaneous",                "opex",       False),
    ("total_fac",   ["TOTAL FACILITY OPEX", "TOTAL FACILITY COSTS"],            "TOTAL FACILITY OPEX",              "facility",   True ),
    ("da",          ["D&A", "Depreciation"],                                     "    D&A",                          "da",         False),
    ("total_costs", ["TOTAL COSTS", "TOTAL DIRECT COSTS"],                      "TOTAL COSTS  (incl. D&A)",         "totals",     True ),
    ("ebitda",      ["EBITDA"],                                                  "EBITDA",                           "profit",     True ),
    ("ebit",        ["EBIT  (", "EBIT ("],                                       "EBIT  (Net Income / Loss)",        "profit",     True ),
]

SCENARIOS = ["BASE", "CONSTRAINED", "STRETCH"]

# Data tab column layout: A=School B=Scenario C=RowKey D=Y1 E=Y2 F=Y3 G=Y4 H=Y5 I=Actuals
DATA_YEAR_COLS = {1: 'D', 2: 'E', 3: 'F', 4: 'G', 5: 'H'}
DATA_ACT_COL   = 'I'   # SY25/26 Actuals column in _Data
DATA_ROW_START = 2
DATA_ROW_END   = 4000  # supports up to ~60 schools × 3 scenarios × 18 rows = 3,240 rows

# School filter panel — top-LEFT sidebar on each scenario tab
# Located in cols A-B (1-2), starts at row 1 (always visible, freeze panes at C3)
# Capacity: 60 school slots (rows 3-62), expandable by bumping FILTER_END.
FILTER_COL_NAME = 1    # col A: school names
FILTER_COL_YN   = 2    # col B: Y/N dropdown toggle
FILTER_HDR_ROW  = 1    # panel title row (row 1, alongside P&L title in cols C-O)
FILTER_SUB_ROW  = 2    # column sub-headers row (row 2, alongside P&L subtitle)
FILTER_START    = 3    # first school row
FILTER_END      = 62   # last slot (rows 3-62 = 60 capacity; add more by increasing this)
FILTER_SUM_ROW  = 63   # summary count row

# P&L column positions
# Layout: A=filter names, B=Y/N, C=separator gap, D=labels, E=actuals,
#         F-I=Q1-Q4, J=Y1, K-N=Y2-Y5, O=5yr total, P=notes
PNL_LABEL_COL = 4    # col D: row labels
PNL_ACT_COL   = 5    # col E: SY25/26 actuals
PNL_Q1_COL    = 6    # col F: Q1
PNL_Q2_COL    = 7    # col G: Q2
PNL_Q3_COL    = 8    # col H: Q3
PNL_Q4_COL    = 9    # col I: Q4
PNL_Y1_COL    = 10   # col J: Year 1 (SY26/27 full year)
PNL_Y2_COL    = 11   # col K: Year 2
PNL_Y3_COL    = 12   # col L: Year 3
PNL_Y4_COL    = 13   # col M: Year 4
PNL_Y5_COL    = 14   # col N: Year 5
PNL_5YR_COL   = 15   # col O: 5-year total
PNL_NOTE_COL  = 16   # col P: notes


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Extract data from each school file
# ─────────────────────────────────────────────────────────────────────────────

def find_row(ws, patterns, max_row=60):
    """Return row number where col A starts with any pattern (case-insensitive startswith).
    Using startswith ensures 'EBIT  (' doesn't match 'EBITDA' and 'Guides' doesn't
    match 'Lead Guides'.
    """
    for r in range(1, max_row + 1):
        v = str(ws.cell(r, 1).value or "").strip().upper()
        for p in patterns:
            ps = p.strip().upper()
            if v == ps or v.startswith(ps):
                return r
    return None


def cell_num(ws, row, col):
    """Read a cell as float; return 0 for None, text, or formula strings."""
    v = ws.cell(row, col).value
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").replace("(", "-").replace(")", ""))
        except ValueError:
            pass
    return 0.0


# ── Kirkland formula evaluator ────────────────────────────────────────────────
# Kirkland uses =SUM(C7:F7), =H14*1.03, =135*65000, =G40+G41 etc.
# Because openpyxl data_only=True can't evaluate formulas without cached values
# we use a mini evaluator that handles the specific patterns Kirkland uses.

def evaluate_kirkland(ws, max_row=55, max_col=12):
    """Return dict {(row, col): numeric_value} for all cells in a Kirkland sheet."""
    import re
    col_n = lambda L: ord(L.upper()) - ord('A') + 1

    raw = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            raw[(r, c)] = v

    cache = {}

    def get(r, c, depth=0):
        if depth > 25: return 0.0
        if (r, c) in cache: return cache[(r, c)]
        v = raw.get((r, c))
        if isinstance(v, (int, float)):
            result = float(v)
        elif isinstance(v, str) and v.startswith("="):
            result = evaluate(v[1:], r, c, depth)
        else:
            result = 0.0
        cache[(r, c)] = result
        return result

    def parse_ref(s):
        m = re.match(r"\$?([A-Z]+)\$?(\d+)", s.strip())
        return (int(m.group(2)), col_n(m.group(1))) if m else None

    def evaluate(expr, br, bc, depth):
        expr = expr.strip()
        # SUM(A1:B3) with optional trailing ±cell
        m = re.match(r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)(.*)", expr, re.I)
        if m:
            r1, c1 = int(m.group(2)), col_n(m.group(1))
            r2, c2 = int(m.group(4)), col_n(m.group(3))
            total = sum(get(r, c, depth + 1) for r in range(r1, r2 + 1)
                        for c in range(c1, c2 + 1))
            rest = m.group(5).strip()
            for part in re.finditer(r"([+\-])\s*([A-Z]+\$?\d+|\d+\.?\d*)", rest, re.I):
                sign = -1 if part.group(1) == "-" else 1
                token = part.group(2)
                ref = parse_ref(token)
                total += sign * (get(*ref, depth + 1) if ref else float(token))
            return total
        # cell*num or num*cell or num*num
        m = re.match(r"([A-Z]+\d+|\d+\.?\d*)\*([A-Z]+\d+|\d+\.?\d*)", expr, re.I)
        if m:
            a, b = m.group(1), m.group(2)
            va = get(*parse_ref(a), depth + 1) if re.match(r"[A-Z]", a, re.I) else float(a)
            vb = get(*parse_ref(b), depth + 1) if re.match(r"[A-Z]", b, re.I) else float(b)
            return va * vb
        # Multi-term: cell±cell±cell…
        terms = re.split(r"(?=[+\-])", expr)
        if len(terms) > 1:
            total = 0.0
            for term in terms:
                term = term.strip()
                if not term: continue
                sign = 1
                if term.startswith("-"):  sign = -1; term = term[1:]
                elif term.startswith("+"): term = term[1:]
                ref = parse_ref(term.strip())
                total += sign * (get(*ref, depth + 1) if ref else 0.0)
            return total
        # Direct cell reference
        ref = parse_ref(expr)
        return get(*ref, depth + 1) if ref else 0.0

    # Evaluate everything
    for key in list(raw.keys()):
        get(*key)
    return cache


def extract_school(school_info):
    """Return dict: {scenario: {row_key: [y1, y2, y3, y4, y5]}}"""
    # Schools with no file (e.g., Fort Worth placeholder) return empty data
    if school_info.get("file") is None:
        return {}
    fpath = os.path.join(BUDGET_DIR, school_info["file"])
    if not os.path.exists(fpath):
        print(f"  ⚠  File not found: {school_info['file']} — skipping")
        return {}

    is_kirkland = "Kirkland" in school_info["file"]

    try:
        # For Kirkland, load with formulas (data_only=False) and evaluate
        wb = openpyxl.load_workbook(fpath, data_only=(not is_kirkland))
    except Exception as e:
        print(f"  ⚠  Could not open {school_info['file']}: {e}")
        return {}

    result = {}
    for scenario in SCENARIOS:
        if scenario not in wb.sheetnames:
            continue
        ws = wb[scenario]

        # For Kirkland: compute all cell values using the mini evaluator
        if is_kirkland:
            cell_cache = evaluate_kirkland(ws)
            def read_val(row, col):
                return cell_cache.get((row, col), 0.0)
        else:
            def read_val(row, col):
                return cell_num(ws, row, col)

        row_data = {}
        row_actuals = {}  # SY25/26 actuals per P&L row (col B)
        for key, patterns, label, section, is_total in PL_ROWS:
            row_num = find_row(ws, patterns)
            if row_num is None:
                row_data[key] = [0.0] * 5
                row_actuals[key] = 0.0
                continue
            # SY25/26 Actuals always in col B (index 2)
            row_actuals[key] = read_val(row_num, 2)

            if key == "enrollment":
                # Enrollment Y1 detection — three file vintages:
                # • New-build (unequal qtrs): Q1-Q4 = annual÷4 (e.g. 6,6,6,7 → sum=25)
                # • New-build (equal qtrs):   Q1-Q4 all equal but sum < Y2×1.5 (e.g. 5,5,5,5 → sum=20)
                # • Plano-style prior build:  Q1-Q4 all equal at annual count (e.g. 19,19,19,19)
                # • Chantilly/col-G prior:    col G has the annual count directly
                g  = read_val(row_num, 7)  # col G (Y1 full year — some prior builds)
                q1 = read_val(row_num, 3)  # col C (Q1)
                q2 = read_val(row_num, 4)  # col D (Q2)
                q3 = read_val(row_num, 5)  # col E (Q3)
                q4 = read_val(row_num, 6)  # col F (Q4)
                y2_peek = read_val(row_num, 8)  # col H — used to detect overcounting
                if g > 0:
                    # Prior-build: col G holds the annual enrollment directly
                    y1 = g
                elif q1 > 0 and q1 == q2 == q3 == q4:
                    # All four quarters equal — could be:
                    #   Plano (19/qtr = 19 annual) OR La Jolla/equal-split (5/qtr = 20 annual)
                    # If summing 4 quarters gives more than 1.5× Y2, it's per-quarter headcount
                    total = q1 + q2 + q3 + q4
                    if y2_peek > 0 and total > y2_peek * 1.5:
                        y1 = q1   # Plano-style: each quarter = annual count
                    else:
                        y1 = total  # Equal-split quarterly fractions → sum to annual
                else:
                    # Unequal quarters: quarterly fractions that sum to annual (most new builds)
                    y1 = q1 + q2 + q3 + q4
                # Some models (e.g. Chantilly) put a dash separator in col H for
                # the enrollment row, pushing Y2+ one column to the right.
                # Detect: if col H = 0 but col I > 0, shift year readings right.
                y2_std = read_val(row_num, 8)    # col H (Y2 in standard format)
                if y2_std == 0 and read_val(row_num, 9) > 0:
                    y2 = read_val(row_num, 9)    # col I
                    y3 = read_val(row_num, 10)   # col J
                    y4 = read_val(row_num, 11)   # col K
                    y5 = read_val(row_num, 12)   # col L
                else:
                    y2 = y2_std                  # col H (normal)
                    y3 = read_val(row_num, 9)    # col I
                    y4 = read_val(row_num, 10)   # col J
                    y5 = read_val(row_num, 11)   # col K
                row_data[key] = [y1, y2, y3, y4, y5]
            else:
                # Year cols: G=7(Y1), H=8(Y2), I=9(Y3), J=10(Y4), K=11(Y5)
                vals = [read_val(row_num, c) for c in [7, 8, 9, 10, 11]]
                row_data[key] = vals
        result[scenario] = {"data": row_data, "actuals": row_actuals}

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Style helpers
# ─────────────────────────────────────────────────────────────────────────────

def sc(ws, row, col, val, fmt=FMT_NUM, bold=False, color=BLACK,
       bg=None, align="right", size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill("solid", fgColor=bg)
    return c


def sec_hdr(ws, row, label, start_col=PNL_LABEL_COL, end_col=PNL_NOTE_COL):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=label)
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Build _Data tab
# ─────────────────────────────────────────────────────────────────────────────

def build_data_tab(ws, all_school_data):
    """Populate hidden _Data tab with one row per school × scenario × P&L row."""
    ws.sheet_properties.tabColor = "007030A0"

    # Headers
    for col, hdr in enumerate(["School", "Scenario", "RowKey", "Y1", "Y2", "Y3", "Y4", "Y5", "Actuals"], 1):
        c = ws.cell(1, col, hdr)
        c.font  = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill  = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for col, w in zip("ABCDEFGHI", [28, 14, 18, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    row = 2
    for school in SCHOOLS:
        sdata = all_school_data.get(school["name"], {})
        for scenario in SCENARIOS:
            sc_block  = sdata.get(scenario, {})
            sc_data   = sc_block.get("data",    {})
            sc_act    = sc_block.get("actuals", {})
            for key, patterns, label, section, is_total in PL_ROWS:
                vals = sc_data.get(key, [0.0, 0.0, 0.0, 0.0, 0.0])
                act  = sc_act.get(key, 0.0)
                ws.cell(row, 1, school["name"]).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row, 2, scenario).alignment      = Alignment(horizontal="center", vertical="center")
                ws.cell(row, 3, key).alignment           = Alignment(horizontal="left", vertical="center")
                for i, v in enumerate(vals[:5]):
                    c = ws.cell(row, 4 + i, v)
                    c.number_format = FMT_NUM
                    c.alignment     = Alignment(horizontal="right", vertical="center")
                # col I = Actuals (SY25/26)
                c = ws.cell(row, 9, act)
                c.number_format = FMT_NUM
                c.alignment     = Alignment(horizontal="right", vertical="center")
                row += 1

    print(f"  _Data tab: {row - 2} rows written ({N_SCHOOLS} schools × 3 scenarios × {len(PL_ROWS)} rows)")


# ─────────────────────────────────────────────────────────────────────────────
# Step 4: Build Config tab
# ─────────────────────────────────────────────────────────────────────────────

def build_config_tab(ws, all_school_data):
    """Config tab: school reference overview (no Y/N filter — filter lives in BASE tab)."""
    ws.sheet_properties.tabColor = "00FFC000"

    # Columns: A=School, B=Location, C=Y1 Enroll, D=Y1 Revenue, E=Y1 EBITDA, F=Y1 EBIT, G=File, H=Assumptions Doc
    for col, w in zip("ABCDEFGH", [28, 26, 14, 18, 18, 18, 16, 16]):
        ws.column_dimensions[col].width = w

    # Row 1: title
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "Alpha Schools SY26/27 — Portfolio Overview  |  Use the BASE tab to filter schools")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=12)
    c.fill      = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2: instructions
    ws.merge_cells("A2:H2")
    c = ws.cell(2, 1, "Reference table — Year 1 BASE figures for all schools.  "
                       "To filter the P&L views, use the Y/N sidebar in the BASE tab.")
    c.font      = Font(name="Calibri", italic=True, color=GRAY_TEXT, size=9)
    c.fill      = PatternFill("solid", fgColor="00F2F2F2")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: column headers
    headers = ["School", "Location", "Y1 Enroll\n(BASE)", "Y1 Revenue\n(BASE)",
               "Y1 EBITDA\n(BASE)", "Y1 EBIT\n(BASE)", "File", "Assumptions\nDoc"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Rows 4+: one per school
    LINK_CLR = "0563C1"   # standard Excel hyperlink blue
    for idx, school in enumerate(SCHOOLS):
        row = CONFIG_DATA_START + idx
        ws.row_dimensions[row].height = 16

        sdata  = all_school_data.get(school["name"], {}).get("BASE", {}).get("data", {})
        enroll = sdata.get("enrollment", [0]*5)[0]
        rev    = sdata.get("total_rev",  [0]*5)[0]
        ebitda = sdata.get("ebitda",     [0]*5)[0]
        ebit   = sdata.get("ebit",       [0]*5)[0]

        sc(ws, row, 1, school["name"],     fmt=FMT_TEXT, align="left", bold=True)
        sc(ws, row, 2, school["location"], fmt=FMT_TEXT, color=GRAY_TEXT, align="left")
        sc(ws, row, 3, enroll,             fmt=FMT_NUM,  align="right")
        sc(ws, row, 4, rev,                fmt=FMT_NUM,  align="right")
        sc(ws, row, 5, ebitda,             fmt=FMT_NUM,  align="right",
           color=("0070AD47" if ebitda >= 0 else "00C00000"))
        sc(ws, row, 6, ebit,               fmt=FMT_NUM,  align="right",
           color=("00000000" if ebit >= 0 else "00C00000"))
        sc(ws, row, 7, school["file"] or school.get("comment","— no model file"),
           fmt=FMT_TEXT, color=GRAY_TEXT, align="left", size=8)

        # Column H: clickable hyperlink to the Assumptions doc on Google Drive
        drive_url = school.get("assumptions_url")
        h_cell = ws.cell(row, 8)
        if drive_url:
            h_cell.value     = f'=HYPERLINK("{drive_url}","Open Assumptions")'
            h_cell.font      = Font(name="Calibri", color=LINK_CLR, underline="single", size=9)
        else:
            h_cell.value = "—"
            h_cell.font  = Font(name="Calibri", color="999999", size=9)
        h_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Placeholder rows for future schools
    for row in range(CONFIG_DATA_START + N_SCHOOLS, CONFIG_DATA_END + 1):
        ws.row_dimensions[row].height = 14

    # Portfolio totals row (Y1, BASE) — uses correct .get("data",{}) accessor
    tot_row = CONFIG_DATA_END + 2
    ws.row_dimensions[tot_row].height = 18
    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    sc(ws, tot_row, 1, "PORTFOLIO TOTAL (all schools, BASE, Y1):", fmt=FMT_TEXT,
       bold=True, align="left", bg=HDR_BG, color=WHITE)
    total_enroll = sum(all_school_data.get(s["name"], {}).get("BASE", {}).get("data", {}).get("enrollment", [0]*5)[0] for s in SCHOOLS)
    total_rev    = sum(all_school_data.get(s["name"], {}).get("BASE", {}).get("data", {}).get("total_rev",  [0]*5)[0] for s in SCHOOLS)
    total_ebitda = sum(all_school_data.get(s["name"], {}).get("BASE", {}).get("data", {}).get("ebitda",     [0]*5)[0] for s in SCHOOLS)
    total_ebit   = sum(all_school_data.get(s["name"], {}).get("BASE", {}).get("data", {}).get("ebit",       [0]*5)[0] for s in SCHOOLS)
    sc(ws, tot_row, 3, total_enroll, fmt=FMT_NUM, bold=True, bg=HDR_BG, color=WHITE)
    sc(ws, tot_row, 4, total_rev,    fmt=FMT_NUM, bold=True, bg=HDR_BG, color=WHITE)
    sc(ws, tot_row, 5, total_ebitda, fmt=FMT_NUM, bold=True, bg=HDR_BG, color=WHITE)
    sc(ws, tot_row, 6, total_ebit,   fmt=FMT_NUM, bold=True, bg=HDR_BG, color=WHITE)


# ─────────────────────────────────────────────────────────────────────────────
# Step 5: Build scenario tabs (BASE / CONSTRAINED / STRETCH)
# ─────────────────────────────────────────────────────────────────────────────

def _filter_clause():
    """The school-filter MATCH clause — references the local tab's sidebar (cols A/B)."""
    fs, fe = FILTER_START, FILTER_END
    nc = get_column_letter(FILTER_COL_NAME)  # A: school names
    yc = get_column_letter(FILTER_COL_YN)    # B: Y/N toggles
    return (
        f'ISNUMBER(MATCH(_Data!$A${DATA_ROW_START}:$A${DATA_ROW_END},'
        f'IF(${yc}${fs}:${yc}${fe}="Y",${nc}${fs}:${nc}${fe},"__none__"),0))'
    )


def sp(scenario, row_key, year_num):
    """SUMPRODUCT formula: aggregates annual data across selected schools."""
    ycol = DATA_YEAR_COLS[year_num]
    return (
        f'=SUMPRODUCT('
        f'(_Data!$B${DATA_ROW_START}:$B${DATA_ROW_END}="{scenario}")*'
        f'(_Data!$C${DATA_ROW_START}:$C${DATA_ROW_END}="{row_key}")*'
        f'{_filter_clause()}*'
        f'_Data!${ycol}${DATA_ROW_START}:${ycol}${DATA_ROW_END})'
    )


def sp_act(scenario, row_key):
    """SUMPRODUCT formula: aggregates SY25/26 actuals across selected schools."""
    return (
        f'=SUMPRODUCT('
        f'(_Data!$B${DATA_ROW_START}:$B${DATA_ROW_END}="{scenario}")*'
        f'(_Data!$C${DATA_ROW_START}:$C${DATA_ROW_END}="{row_key}")*'
        f'{_filter_clause()}*'
        f'_Data!${DATA_ACT_COL}${DATA_ROW_START}:${DATA_ACT_COL}${DATA_ROW_END})'
    )


def build_scenario_tab(ws, scenario, tab_color, school_list, is_base=False, school_status=None):
    """Build a consolidated P&L scenario tab with SUMPRODUCT formulas and top-left filter sidebar.

    is_base=True  → interactive filter (BASE tab only): Y/N dropdowns, data validation
    is_base=False → read-only mirror (CONSTRAINED/STRETCH): col B = =BASE!$B{row}
    school_status → dict {school_name: "live" | "launching"} for col A colour coding
    """
    if school_status is None:
        school_status = {}

    # Status colours for col A school names
    LIVE_TXT      = "004472C4"   # Excel standard blue  → already operating
    LAUNCH_TXT    = "00ED7D31"   # warm orange           → opening this year

    # Y/N colour scheme — very subtle: light fill + dark matching text (Excel-standard feel)
    FILTER_Y_BG   = "00E2EFDA"   # light sage green
    FILTER_N_BG   = "00FFDCDC"   # light blush pink
    FILTER_Y_TXT  = "00375623"   # dark green text
    FILTER_N_TXT  = "009C0006"   # dark red text
    ws.sheet_properties.tabColor = tab_color

    # Column widths
    # A=filter names (narrow), B=Y/N (narrow), C=separator gap, D=labels,
    # E=actuals, F-I=Q1-Q4, J-N=Y1-Y5, O=5yr, P=notes
    ws.column_dimensions["A"].width = 17    # school names — narrow but readable
    ws.column_dimensions["B"].width = 4.5  # Y/N toggle — just wide enough for the letter
    ws.column_dimensions["C"].width = 1.5  # visual separator gap between filter and P&L
    ws.column_dimensions["D"].width = 34   # P&L row labels
    ws.column_dimensions["E"].width = 12   # SY25/26 Actuals
    for col in ["F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 11  # Q1-Q4
    for col in ["J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 13  # Y1-Y5
    ws.column_dimensions["O"].width = 14   # 5-Year Total
    ws.column_dimensions["P"].width = 32   # Notes

    # Row heights
    row_heights = {
        1: 22, 2: 18, 3: 36, 4: 18, 5: 5,
        6: 18, 7: 16, 8: 16, 9: 16, 10: 16, 11: 18, 12: 5,
        13: 18, 14: 16, 15: 16, 16: 16, 17: 16, 18: 18, 19: 5,
        20: 18, 21: 16, 22: 16, 23: 16, 24: 5,
        25: 18, 26: 18, 27: 5,
        28: 18, 29: 16, 30: 5,
        31: 20, 32: 5,
        33: 18, 34: 18,
        FILTER_HDR_ROW: 18,
    }
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h
    for r in range(FILTER_START, FILTER_END + 1):
        ws.row_dimensions[r].height = 16

    # ── Row 1: Filter title (A-B) + separator (C) + P&L title (D-P) ─────
    # Filter panel header
    ws.merge_cells("A1:B1")
    filter_title = "▼  SCHOOL FILTER" if is_base else "↔  FILTER  (mirrors BASE tab)"
    filter_hdr_bg = TITLE_BG if is_base else HDR_BG
    c = ws.cell(1, 1, filter_title)
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
    c.fill      = PatternFill("solid", fgColor=filter_hdr_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # Col C: white separator — pure white across all rows, acts as visual gap
    for _r in range(1, FILTER_END + 5):
        ws.cell(_r, 3).fill = PatternFill("solid", fgColor=WHITE)
    # P&L title
    pnl_last_col = get_column_letter(PNL_NOTE_COL)
    pnl_first_col = get_column_letter(PNL_LABEL_COL)
    ws.merge_cells(f"{pnl_first_col}1:{pnl_last_col}1")
    c = ws.cell(1, PNL_LABEL_COL, f"Alpha Schools Portfolio  |  {scenario}  |  SY 2026/27 Consolidated Budget")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 2: Filter sub-headers (A-B) + separator (C) + live subtitle (D-P) ──
    yc = get_column_letter(FILTER_COL_YN)
    # A2: "School" sub-header (doubles as colour legend)
    c = ws.cell(2, 1, "School  ● live  ◌ launching")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=7)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    # B2: "In?" sub-header (compact)
    c = ws.cell(2, 2, "In?")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=8)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # C2: already set to white by the separator loop above — no additional styling
    # D2:P2: live selected-schools count
    ws.merge_cells(f"{pnl_first_col}2:{pnl_last_col}2")
    filter_hint = "← Toggle Y/N in left sidebar to filter" if is_base else "← Filter is set in BASE tab (read-only here)"
    selected_formula = (
        f'=COUNTIF(${yc}${FILTER_START}:${yc}${FILTER_END},"Y")'
        f'&" of {N_SCHOOLS} schools included  ·  '
        f'{filter_hint}  ·  '
        f'Re-run refresh_consolidation.py after editing school files"'
    )
    c = ws.cell(2, PNL_LABEL_COL, selected_formula)
    c.font      = Font(name="Calibri", italic=True, color=GRAY_TEXT, size=9)
    c.fill      = PatternFill("solid", fgColor=ADDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Column headers (D-P); col C stays white (set by separator loop) ──
    for i, h in enumerate(
        ["", "SY25/26\nActuals", "Q1\n(est.)", "Q2\n(est.)", "Q3\n(est.)", "Q4\n(est.)",
         "SY26/27\nFull Year", "Y2\nSY27/28", "Y3\nSY28/29", "Y4\nSY29/30",
         "Y5\nSY30/31", "5-Year\nTotal", "Notes"]
    ):
        c = ws.cell(3, PNL_LABEL_COL + i, h)
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center" if i > 0 else "left",
                                vertical="center", wrap_text=True)

    # ── data_row helper ───────────────────────────────────────────────────
    # Layout: D=label, E=actuals, F-I=Q1-Q4, J=Y1, K-N=Y2-Y5, O=5yr, P=notes
    # ALL data columns (E through O) use the same text colour — no grey distinction
    # between actuals/quarterly estimates and annual forecast years.
    def data_row(pl_row, ws_row, note="", bold=False, bg=None, txt=BLACK):
        key, patterns, label, section, is_total = pl_row
        # On dark backgrounds force white; otherwise use the passed-in txt colour
        num_txt  = WHITE if bg in (HDR_BG, DARK_BG) else txt
        note_txt = "00AAAAAA" if bg in (HDR_BG, DARK_BG) else GRAY_TEXT

        # Col D (4): label
        c = ws.cell(ws_row, PNL_LABEL_COL, label)
        c.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.number_format = FMT_TEXT
        if bg: c.fill = PatternFill("solid", fgColor=bg)

        # Col E (5): SY25/26 Actuals — same colour treatment as forecast cols
        act_f = sp_act(scenario, key)
        c2 = ws.cell(ws_row, PNL_ACT_COL, act_f)
        c2.number_format = FMT_NUM
        c2.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
        c2.alignment = Alignment(horizontal="right", vertical="center")
        if bg: c2.fill = PatternFill("solid", fgColor=bg)

        # Col J (10): Y1 full-year SUMPRODUCT
        f_y1 = sp(scenario, key, 1)
        c_y1 = ws.cell(ws_row, PNL_Y1_COL, f_y1)
        c_y1.number_format = FMT_NUM
        c_y1.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
        c_y1.alignment = Alignment(horizontal="right", vertical="center")
        if bg: c_y1.fill = PatternFill("solid", fgColor=bg)

        # Cols F-I (6-9): quarterly estimate = ROUND(Y1/4, 0) — same colour as Y1
        y1_addr = f"J{ws_row}"
        for col_idx in [PNL_Q1_COL, PNL_Q2_COL, PNL_Q3_COL, PNL_Q4_COL]:
            if key == "enrollment" and col_idx == PNL_Q4_COL:
                formula = f"={y1_addr}-3*ROUND({y1_addr}/4,0)"
            else:
                formula = f"=ROUND({y1_addr}/4,0)"
            c = ws.cell(ws_row, col_idx, formula)
            c.number_format = FMT_NUM
            c.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if bg: c.fill = PatternFill("solid", fgColor=bg)

        # Cols K-N (11-14): Y2-Y5 SUMPRODUCT
        for yr, col_idx in zip([2, 3, 4, 5], [PNL_Y2_COL, PNL_Y3_COL, PNL_Y4_COL, PNL_Y5_COL]):
            f = sp(scenario, key, yr)
            c = ws.cell(ws_row, col_idx, f)
            c.number_format = FMT_NUM
            c.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if bg: c.fill = PatternFill("solid", fgColor=bg)

        # Col O (15): 5-Year Total  (J:N = Y1 through Y5)
        f5 = f"=SUM(J{ws_row}:N{ws_row})" if key != "enrollment" else f"=J{ws_row}"
        c = ws.cell(ws_row, PNL_5YR_COL, f5)
        c.number_format = FMT_NUM
        c.font      = Font(name="Calibri", bold=bold, color=num_txt, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")
        if bg: c.fill = PatternFill("solid", fgColor=bg)

        # Col P (16): Notes — kept subtle (lighter text)
        c = ws.cell(ws_row, PNL_NOTE_COL, note)
        c.font      = Font(name="Calibri", color=note_txt, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.number_format = FMT_TEXT
        if bg: c.fill = PatternFill("solid", fgColor=bg)

    # ── Row 4: Enrollment ─────────────────────────────────────────────────
    data_row(next(r for r in PL_ROWS if r[0] == "enrollment"), 4,
             note="Enrolled students (Y1 total, quarterly = Y1÷4 estimate)", bold=True, bg=ENROLL_BG)

    # ── REVENUE (rows 6–11) ───────────────────────────────────────────────
    sec_hdr(ws, 6, "REVENUE")
    data_row(next(r for r in PL_ROWS if r[0] == "core_rev"),  7,
             note="Sum of (enrollment × tuition) across selected schools")
    data_row(next(r for r in PL_ROWS if r[0] == "ah_rev"),   8,
             note="After-hours programs — schools with AH model only")
    data_row(next(r for r in PL_ROWS if r[0] == "ss_rev"),   9,
             note="Summer school — schools with SS model only")
    sc(ws, 10, PNL_LABEL_COL, "    Financial Aid / Discount", fmt=FMT_TEXT, align="left")
    for col_idx in [PNL_ACT_COL, PNL_Q1_COL, PNL_Q2_COL, PNL_Q3_COL, PNL_Q4_COL,
                    PNL_Y1_COL, PNL_Y2_COL, PNL_Y3_COL, PNL_Y4_COL, PNL_Y5_COL, PNL_5YR_COL]:
        ws.cell(10, col_idx, 0).number_format = FMT_NUM
        ws.cell(10, col_idx).font = Font(name="Calibri", color=BLACK, size=10)
        ws.cell(10, col_idx).alignment = Alignment(horizontal="right", vertical="center")
    sc(ws, 10, PNL_NOTE_COL, "None modelled", fmt=FMT_TEXT, color=GRAY_TEXT, align="left")
    # TOTAL REVENUE: dark blue bg → white text throughout
    data_row(next(r for r in PL_ROWS if r[0] == "total_rev"), 11, "—",
             bold=True, bg=HDR_BG, txt=WHITE)

    # ── HEADCOUNT (rows 13–18) ────────────────────────────────────────────
    sec_hdr(ws, 13, "HEADCOUNT COSTS  (115% loaded salaries  ·  mixed tuition tiers)")
    data_row(next(r for r in PL_ROWS if r[0] == "lead_guides"), 14,
             note="1–2 lead guides per school depending on enrollment")
    data_row(next(r for r in PL_ROWS if r[0] == "hos"),         15,
             note="Head of School present at Chicago & Santa Barbara only")
    data_row(next(r for r in PL_ROWS if r[0] == "guides"),      16)
    data_row(next(r for r in PL_ROWS if r[0] == "admin"),       17)
    data_row(next(r for r in PL_ROWS if r[0] == "total_hc"),    18, "—", bold=True, bg=YELLOW_BG)

    # ── PROGRAMS, TIMEBACK & MISC (rows 20–23) ────────────────────────────
    sec_hdr(ws, 20, "PROGRAMS, TIMEBACK & MISC")
    data_row(next(r for r in PL_ROWS if r[0] == "timeback"), 21,
             note="20% of core tuition revenue  ·  expensed, not netted")
    data_row(next(r for r in PL_ROWS if r[0] == "programs"), 22)
    data_row(next(r for r in PL_ROWS if r[0] == "misc"),     23)

    # ── FACILITY (rows 25–26) ─────────────────────────────────────────────
    sec_hdr(ws, 25, "FACILITY & LEASE COSTS  (portfolio aggregate  ·  line detail in individual school files)")
    data_row(next(r for r in PL_ROWS if r[0] == "total_fac"), 26,
             note="Sum of each school's total facility cost; detail in individual files",
             bold=True, bg=YELLOW_BG)

    # ── D&A (rows 28–29) ──────────────────────────────────────────────────
    sec_hdr(ws, 28, "DEPRECIATION & AMORTIZATION")
    data_row(next(r for r in PL_ROWS if r[0] == "da"), 29,
             note="Straight-line D&A; varies by school CapEx and term")

    # ── TOTAL COSTS (row 31) — dark bg → white text ───────────────────────
    data_row(next(r for r in PL_ROWS if r[0] == "total_costs"), 31,
             "—", bold=True, bg=DARK_BG, txt=WHITE)

    # ── EBITDA (row 33) ───────────────────────────────────────────────────
    data_row(next(r for r in PL_ROWS if r[0] == "ebitda"), 33, "—", bold=True, bg=GREEN_BG)
    # ── EBIT (row 34) ─────────────────────────────────────────────────────
    data_row(next(r for r in PL_ROWS if r[0] == "ebit"),   34, "—", bold=True, bg=YELLOW_BG)

    # ── LEFT FILTER SIDEBAR (cols A-B, rows 3-25) ────────────────────────
    # Rows 1-2 of the sidebar (title + sub-headers) are already written
    # above alongside the P&L title/subtitle.  Here we fill the school rows,
    # placeholder rows, summary row, data validation, and conditional formatting.

    nc = get_column_letter(FILTER_COL_NAME)   # A
    yc = get_column_letter(FILTER_COL_YN)     # B

    # ── School rows (rows 3 → FILTER_END) ────────────────────────────────
    for idx, school in enumerate(school_list):
        r   = FILTER_START + idx
        alt = "00F7F7F7" if idx % 2 == 0 else "00FFFFFF"
        # School name cell (col A) — colour indicates live vs launching
        status   = school_status.get(school["name"], "launching")
        name_txt = LIVE_TXT if status == "live" else LAUNCH_TXT
        c_n = ws.cell(r, FILTER_COL_NAME, school["name"])
        c_n.font      = Font(name="Calibri", size=9, bold=(status == "live"), color=name_txt)
        c_n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_n.fill      = PatternFill("solid", fgColor=alt)
        # Y/N toggle cell (col B):
        #   BASE tab → hardcoded "Y" with data validation (user edits here)
        #   Other tabs → formula =BASE!$B{r} (mirrors BASE selection, read-only)
        if is_base:
            c_y = ws.cell(r, FILTER_COL_YN, "Y")
        else:
            c_y = ws.cell(r, FILTER_COL_YN, f"=BASE!$B{r}")
        c_y.font      = Font(name="Calibri", bold=True, size=9, color=FILTER_Y_TXT)
        c_y.alignment = Alignment(horizontal="center", vertical="center")
        c_y.fill      = PatternFill("solid", fgColor=FILTER_Y_BG)  # green default; CF overrides

    # ── Placeholder rows for future schools ───────────────────────────────
    # col B is EMPTY (not "Y") so COUNTIF("Y") stays accurate — only real
    # school rows count.  When a new school is added to SCHOOLS and the script
    # is re-run, that row is promoted from placeholder to an active school row.
    for r in range(FILTER_START + len(school_list), FILTER_END + 1):
        c_n = ws.cell(r, FILTER_COL_NAME, "")
        c_n.fill      = PatternFill("solid", fgColor="00F2F2F2")
        c_n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_n.font      = Font(name="Calibri", size=9, color=GRAY_TEXT)
        if is_base:
            c_y = ws.cell(r, FILTER_COL_YN, "")   # empty — not counted by COUNTIF
        else:
            c_y = ws.cell(r, FILTER_COL_YN, f"=BASE!$B{r}")   # mirrors BASE (empty)
        c_y.fill      = PatternFill("solid", fgColor="00F2F2F2")
        c_y.alignment = Alignment(horizontal="center", vertical="center")
        c_y.font      = Font(name="Calibri", size=9, color=GRAY_TEXT)

    # ── Summary count row ─────────────────────────────────────────────────
    ws.row_dimensions[FILTER_SUM_ROW].height = 16
    ws.merge_cells(f"{nc}{FILTER_SUM_ROW}:{get_column_letter(FILTER_COL_YN)}{FILTER_SUM_ROW}")
    c_s = ws.cell(FILTER_SUM_ROW, FILTER_COL_NAME,
        f'=COUNTIF(${yc}${FILTER_START}:${yc}${FILTER_END},"Y")&" / {N_SCHOOLS} selected"')
    c_s.font      = Font(name="Calibri", italic=True, bold=True, color="00375623", size=9)
    c_s.alignment = Alignment(horizontal="center", vertical="center")
    c_s.fill      = PatternFill("solid", fgColor="00E8F5E9")

    yn_range = f"{yc}{FILTER_START}:{yc}{FILTER_END}"

    # ── Data validation: only on BASE tab (dropdowns for toggling Y / N) ──
    if is_base:
        dv = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=False,
            showDropDown=False,   # False = arrow IS visible
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error='Please enter Y (include) or N (exclude).',
            showInputMessage=True,
            promptTitle="Include school?",
            prompt="Y = include in portfolio totals\nN = exclude from totals"
        )
        dv.sqref = yn_range
        ws.add_data_validation(dv)

    # ── Conditional formatting: Y → vivid green, N → vivid red (all tabs) ─
    # Same rules work for mirror tabs since =BASE!$B{r} evaluates to "Y"/"N".
    active_y_fill = PatternFill("solid", fgColor=FILTER_Y_BG)
    active_n_fill = PatternFill("solid", fgColor=FILTER_N_BG)
    active_y_font = Font(name="Calibri", bold=True, color=FILTER_Y_TXT, size=9)
    active_n_font = Font(name="Calibri", bold=True, color=FILTER_N_TXT, size=9)

    ws.conditional_formatting.add(
        yn_range,
        CellIsRule(operator="equal", formula=['"Y"'],
                   fill=active_y_fill, font=active_y_font)
    )
    ws.conditional_formatting.add(
        yn_range,
        CellIsRule(operator="equal", formula=['"N"'],
                   fill=active_n_fill, font=active_n_font)
    )

    # ── Freeze: cols A-C (filter + separator) + rows 1-2 (title + subtitle) always visible
    ws.freeze_panes = "D3"


# ─────────────────────────────────────────────────────────────────────────────
# Step 6: Build Summary tab
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_tab(ws, all_school_data, school_status=None):
    """Side-by-side Y1 snapshot — Option C layout.

    • Filter sidebar A-B (mirrors BASE selection; read-only like CONSTRAINED/STRETCH)
    • Col C: white separator
    • Data cols D-P: School, Location, then numeric columns
    • FormulaRule CF: rows where B="N" are greyed out across D-P
    • Portfolio Total: SUMPRODUCT formulas (live, respects the BASE filter)
    • Freeze panes at D3
    """
    if school_status is None:
        school_status = {}

    LIVE_TXT     = "004472C4"   # blue  → already operating
    LAUNCH_TXT   = "00ED7D31"   # orange → opening this year

    FILTER_Y_BG  = "00E2EFDA"   # light sage green
    FILTER_N_BG  = "00FFDCDC"   # light blush pink
    FILTER_Y_TXT = "00375623"   # dark green text
    FILTER_N_TXT = "009C0006"   # dark red text

    ws.sheet_properties.tabColor = "00FFBF00"

    # Column layout: A=filter names, B=Y/N mirror, C=separator,
    # D=school name, E=location,
    # F=Y1EnrollBASE, G=Y1EnrollSTRETCH,
    # H=Y1RevBASE,    I=Y1RevSTRETCH,
    # J=Y1EBITDABASE, K=Y1EBITDASTRETCH,
    # L=Y1EBITBASE,   M=Y1EBITSTRETCH,
    # N=EBITMargin%,  O=5yrRevBASE,  P=5yrEBITBASE
    SUM_SCHOOL_COL = 4   # col D
    SUM_LOC_COL    = 5   # col E
    SUM_LAST_COL   = 16  # col P

    for col, w in zip(
        ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"],
        [17,  4.5, 1.5, 26, 18, 11, 11, 14, 14, 14, 14, 14, 14, 10, 14, 14]
    ):
        ws.column_dimensions[col].width = w

    # Col C: white separator across all rows
    for _r in range(1, FILTER_END + 5):
        ws.cell(_r, 3).fill = PatternFill("solid", fgColor=WHITE)

    # ── Row 1 ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    # A1:B1 — filter panel title (mirrors BASE)
    ws.merge_cells("A1:B1")
    c = ws.cell(1, 1, "↔  FILTER  (mirrors BASE tab)")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    # D1:P1 — summary title
    pnl_first = get_column_letter(SUM_SCHOOL_COL)
    pnl_last  = get_column_letter(SUM_LAST_COL)
    ws.merge_cells(f"{pnl_first}1:{pnl_last}1")
    c = ws.cell(1, SUM_SCHOOL_COL, "Alpha Schools SY26/27  |  Portfolio Summary  |  Year 1 (SY26/27) Snapshot")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 2: sub-headers ────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 32
    c = ws.cell(2, 1, "School  ● live  ◌ launching")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=7)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c = ws.cell(2, 2, "In?")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=8)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # D2:P2 — column headers (data table)
    headers = ["School", "Location",
               "Y1 Enroll\nBASE",    "Y1 Enroll\nSTRETCH",
               "Y1 Revenue\nBASE",   "Y1 Revenue\nSTRETCH",
               "Y1 EBITDA\nBASE",    "Y1 EBITDA\nSTRETCH",
               "Y1 EBIT\nBASE",      "Y1 EBIT\nSTRETCH",
               "EBIT Margin\nBASE",
               "5yr Revenue\nBASE",  "5yr EBIT\nBASE"]
    for i, h in enumerate(headers):
        col_idx = SUM_SCHOOL_COL + i
        c = ws.cell(2, col_idx, h)
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="left" if i < 2 else "center",
                                vertical="center", wrap_text=True)

    # ── School rows (rows 3 → 3+N_SCHOOLS-1) ─────────────────────────────────
    last_school_row = FILTER_START + N_SCHOOLS - 1  # row 21 for 19 schools
    for idx, school in enumerate(SCHOOLS):
        row = FILTER_START + idx
        ws.row_dimensions[row].height = 15

        sdata   = all_school_data.get(school["name"], {})
        base    = sdata.get("BASE",    {}).get("data", {})
        stretch = sdata.get("STRETCH", {}).get("data", {})

        b_enroll  = base.get("enrollment", [0]*5)[0]
        s_enroll  = stretch.get("enrollment", [0]*5)[0]
        b_rev     = base.get("total_rev",   [0]*5)[0]
        s_rev     = stretch.get("total_rev", [0]*5)[0]
        b_ebitda  = base.get("ebitda",      [0]*5)[0]
        s_ebitda  = stretch.get("ebitda",    [0]*5)[0]
        b_ebit    = base.get("ebit",        [0]*5)[0]
        s_ebit    = stretch.get("ebit",      [0]*5)[0]
        b_margin  = (b_ebit / b_rev) if b_rev else 0
        b_5yr_rev = sum(base.get("total_rev", [0]*5))
        b_5yr_ebit= sum(base.get("ebit",      [0]*5))

        row_bg = "00FAFAFA" if idx % 2 == 0 else None

        sc(ws, row, 4,  school["name"],     fmt=FMT_TEXT, align="left",   bg=row_bg, bold=True)
        sc(ws, row, 5,  school["location"], fmt=FMT_TEXT, align="left",   bg=row_bg, color=GRAY_TEXT)
        sc(ws, row, 6,  b_enroll,  fmt=FMT_NUM, align="center", bg=row_bg)
        sc(ws, row, 7,  s_enroll,  fmt=FMT_NUM, align="center", bg=row_bg, color=GRAY_TEXT)
        sc(ws, row, 8,  b_rev,     fmt=FMT_NUM, bg=row_bg)
        sc(ws, row, 9,  s_rev,     fmt=FMT_NUM, bg=row_bg, color=GRAY_TEXT)
        sc(ws, row, 10, b_ebitda,  fmt=FMT_NUM, bg=row_bg,
           color=("0070AD47" if b_ebitda >= 0 else "00C00000"))
        sc(ws, row, 11, s_ebitda,  fmt=FMT_NUM, bg=row_bg,
           color=("0070AD47" if s_ebitda >= 0 else "00C00000"))
        sc(ws, row, 12, b_ebit,    fmt=FMT_NUM, bg=row_bg,
           color=("00000000" if b_ebit >= 0 else "00C00000"))
        sc(ws, row, 13, s_ebit,    fmt=FMT_NUM, bg=row_bg,
           color=("00000000" if s_ebit >= 0 else "00C00000"))

        marg_str = f"{b_margin*100:.1f}%"
        c = ws.cell(row, 14, marg_str)
        c.font = Font(name="Calibri", size=10,
                      color=("0070AD47" if b_margin >= 0 else "00C00000"), bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if row_bg: c.fill = PatternFill("solid", fgColor=row_bg)

        sc(ws, row, 15, b_5yr_rev,  fmt=FMT_NUM, bg=row_bg)
        sc(ws, row, 16, b_5yr_ebit, fmt=FMT_NUM, bg=row_bg,
           color=("00000000" if b_5yr_ebit >= 0 else "00C00000"))

        # If school has a comment (e.g., placeholder with no data yet), write it
        note = school.get("comment", "")
        if note:
            # overwrite col F-P with the note spanning cols F-P
            ws.merge_cells(f"F{row}:P{row}")
            c_note = ws.cell(row, 6, f"⚠  {note}")
            c_note.font      = Font(name="Calibri", size=9, italic=True, color="00999999")
            c_note.alignment = Alignment(horizontal="left", vertical="center")
            if row_bg:
                c_note.fill = PatternFill("solid", fgColor=row_bg)

    # ── FormulaRule CF: grey-out rows excluded from filter (B="N") ────────────
    # Applied to D3:P{last school row}; =$B3="N" anchors col B, row moves with range.
    # CF overrides static font/fill when condition is true.
    grey_fill = PatternFill("solid", fgColor="00D9D9D9")
    grey_font = Font(name="Calibri", color="00999999", size=10)
    ws.conditional_formatting.add(
        f"D{FILTER_START}:P{last_school_row}",
        FormulaRule(formula=[f'=$B{FILTER_START}="N"'], fill=grey_fill, font=grey_font)
    )

    # ── Portfolio Total row (SUMPRODUCT — live, respects BASE filter) ─────────
    # Placed 2 rows below last school row for visual separation.
    tot_row = last_school_row + 2
    ws.row_dimensions[tot_row].height = 18

    ws.merge_cells(f"D{tot_row}:E{tot_row}")
    c = ws.cell(tot_row, 4, "PORTFOLIO TOTAL  (filtered schools, BASE / STRETCH, Y1)")
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Numeric totals via sp() — the SUMPRODUCT references this tab's A/B sidebar
    # (which mirrors BASE), so the total stays in sync with the filter.
    sp_cols = [
        (6,  sp("BASE",    "enrollment", 1), FMT_NUM, "center"),
        (7,  sp("STRETCH", "enrollment", 1), FMT_NUM, "center"),
        (8,  sp("BASE",    "total_rev",  1), FMT_NUM, "right"),
        (9,  sp("STRETCH", "total_rev",  1), FMT_NUM, "right"),
        (10, sp("BASE",    "ebitda",     1), FMT_NUM, "right"),
        (11, sp("STRETCH", "ebitda",     1), FMT_NUM, "right"),
        (12, sp("BASE",    "ebit",       1), FMT_NUM, "right"),
        (13, sp("STRETCH", "ebit",       1), FMT_NUM, "right"),
    ]
    for col_idx, formula, fmt, align in sp_cols:
        c = ws.cell(tot_row, col_idx, formula)
        c.number_format = fmt
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.fill      = PatternFill("solid", fgColor=HDR_BG)

    # EBIT margin% = EBIT / Revenue (formula referencing the cells just written)
    rev_cell  = f"{get_column_letter(8)}{tot_row}"
    ebit_cell = f"{get_column_letter(12)}{tot_row}"
    c = ws.cell(tot_row, 14, f"=IF({rev_cell}<>0,{ebit_cell}/{rev_cell},0)")
    c.number_format = "0.0%"
    c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=HDR_BG)

    # 5-year totals: sum of 5 annual SUMPRODUCT formulas
    def sp5yr(scenario, row_key):
        """Sum of sp(scenario, row_key, 1..5) — a single Excel formula."""
        parts = [sp(scenario, row_key, yr)[1:] for yr in [1, 2, 3, 4, 5]]
        return "=" + "+".join(f"({p})" for p in parts)

    for col_idx, scenario, row_key in [(15, "BASE", "total_rev"), (16, "BASE", "ebit")]:
        c = ws.cell(tot_row, col_idx, sp5yr(scenario, row_key))
        c.number_format = FMT_NUM
        c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill      = PatternFill("solid", fgColor=HDR_BG)

    # ── Left filter sidebar (A-B, mirrors BASE — same pattern as CONSTRAINED/STRETCH) ─
    nc = get_column_letter(FILTER_COL_NAME)  # A
    yc = get_column_letter(FILTER_COL_YN)   # B

    for idx, school in enumerate(SCHOOLS):
        r      = FILTER_START + idx
        alt    = "00F7F7F7" if idx % 2 == 0 else "00FFFFFF"
        status = school_status.get(school["name"], "launching")
        name_txt = LIVE_TXT if status == "live" else LAUNCH_TXT
        c_n = ws.cell(r, FILTER_COL_NAME, school["name"])
        c_n.font      = Font(name="Calibri", size=9, bold=(status == "live"), color=name_txt)
        c_n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_n.fill      = PatternFill("solid", fgColor=alt)
        c_y = ws.cell(r, FILTER_COL_YN, f"=BASE!$B{r}")
        c_y.font      = Font(name="Calibri", bold=True, size=9, color=FILTER_Y_TXT)
        c_y.alignment = Alignment(horizontal="center", vertical="center")
        c_y.fill      = PatternFill("solid", fgColor=FILTER_Y_BG)

    # Placeholder rows for future schools
    for r in range(FILTER_START + N_SCHOOLS, FILTER_END + 1):
        c_n = ws.cell(r, FILTER_COL_NAME, "")
        c_n.fill      = PatternFill("solid", fgColor="00F2F2F2")
        c_n.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_n.font      = Font(name="Calibri", size=9, color=GRAY_TEXT)
        c_y = ws.cell(r, FILTER_COL_YN, f"=BASE!$B{r}")
        c_y.fill      = PatternFill("solid", fgColor="00F2F2F2")
        c_y.alignment = Alignment(horizontal="center", vertical="center")
        c_y.font      = Font(name="Calibri", size=9, color=GRAY_TEXT)

    # Summary count row
    ws.row_dimensions[FILTER_SUM_ROW].height = 16
    ws.merge_cells(f"{nc}{FILTER_SUM_ROW}:{yc}{FILTER_SUM_ROW}")
    c_s = ws.cell(FILTER_SUM_ROW, FILTER_COL_NAME,
        f'=COUNTIF(${yc}${FILTER_START}:${yc}${FILTER_END},"Y")&" / {N_SCHOOLS} selected"')
    c_s.font      = Font(name="Calibri", italic=True, bold=True, color="00375623", size=9)
    c_s.alignment = Alignment(horizontal="center", vertical="center")
    c_s.fill      = PatternFill("solid", fgColor="00E8F5E9")

    # CF for Y/N toggle colours in sidebar (same as scenario tabs)
    yn_range = f"{yc}{FILTER_START}:{yc}{FILTER_END}"
    ws.conditional_formatting.add(
        yn_range,
        CellIsRule(operator="equal", formula=['"Y"'],
                   fill=PatternFill("solid", fgColor=FILTER_Y_BG),
                   font=Font(name="Calibri", bold=True, color=FILTER_Y_TXT, size=9))
    )
    ws.conditional_formatting.add(
        yn_range,
        CellIsRule(operator="equal", formula=['"N"'],
                   fill=PatternFill("solid", fgColor=FILTER_N_BG),
                   font=Font(name="Calibri", bold=True, color=FILTER_N_TXT, size=9))
    )

    ws.freeze_panes = "D3"


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n╔══════════════════════════════════════════════════════════════╗")
    print(  "║  Alpha Schools — SY26/27 Consolidation Builder               ║")
    print(  "╚══════════════════════════════════════════════════════════════╝")
    print(f"\n  Reading {N_SCHOOLS} school files from: {BUDGET_DIR}\n")

    # 1. Extract data from all school files
    all_school_data = {}
    for school in SCHOOLS:
        print(f"  Reading {school['name']} ({school['file'] or 'placeholder — no file'}) ...", end=" ", flush=True)
        data = extract_school(school)
        all_school_data[school["name"]] = data
        if data:
            base_rev = data.get("BASE", {}).get("data", {}).get("total_rev", [0])[0]
            print(f"✓  (BASE Y1 Rev: ${base_rev:,.0f})")
        else:
            print("⚠  (no data)")

    # 1b. Determine live vs launching status for each school
    # "Live" = school had non-zero SY25/26 actuals in their budget file
    # "Launching" = no actuals yet (opening for the first time in SY26/27)
    school_status = {}
    for school in SCHOOLS:
        act = (all_school_data.get(school["name"], {})
               .get("BASE", {}).get("actuals", {}).get("total_rev", 0.0))
        school_status[school["name"]] = "live" if act != 0 else "launching"
    # Hardcode confirmed live schools (some models may not carry actuals in file)
    for _live in ["Alpha Chantilly", "Alpha Charlotte", "Alpha Palo Alto",
                  "Alpha Boston", "Alpha New York", "Alpha Kirkland", "Alpha Plano",
                  "Alpha Houston"]:
        school_status[_live] = "live"
    live_count     = sum(1 for v in school_status.values() if v == "live")
    launching_count = N_SCHOOLS - live_count
    print(f"\n  School status: {live_count} live (blue)  ·  {launching_count} launching (orange)")

    # 2. Build workbook
    print("\n  Building consolidation workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Tab order: Config, BASE, CONSTRAINED, STRETCH, Summary, _Data (hidden)
    ws_config = wb.create_sheet("Config")
    ws_base   = wb.create_sheet("BASE")
    ws_con    = wb.create_sheet("CONSTRAINED")
    ws_str    = wb.create_sheet("STRETCH")
    ws_sum    = wb.create_sheet("Summary")
    ws_data   = wb.create_sheet("_Data")

    # Build each tab
    print("  Building Config tab ...", end=" ", flush=True)
    build_config_tab(ws_config, all_school_data)
    print("✓")

    tab_colors = {"BASE": "004472C4", "CONSTRAINED": "00ED7D31", "STRETCH": "0070AD47"}
    for scenario, ws in [("BASE", ws_base), ("CONSTRAINED", ws_con), ("STRETCH", ws_str)]:
        print(f"  Building {scenario} tab ...", end=" ", flush=True)
        build_scenario_tab(ws, scenario, tab_colors[scenario], SCHOOLS,
                           is_base=(scenario == "BASE"),
                           school_status=school_status)
        print("✓")

    print("  Building Summary tab ...", end=" ", flush=True)
    build_summary_tab(ws_sum, all_school_data, school_status=school_status)
    print("✓")

    print("  Building _Data tab ...", end=" ", flush=True)
    build_data_tab(ws_data, all_school_data)
    ws_data.sheet_state = "hidden"
    print("✓")

    # 3. Save
    wb.save(OUT_FILE)
    print(f"\n  ✅  Saved:  {OUT_FILE}")
    print(f"\n  How to use:")
    print(f"    • Open Alpha_Consolidation_SY2627.xlsx in Excel")
    print(f"    • BASE / CONSTRAINED / STRETCH tabs show the aggregated P&L")
    print(f"    • Filter panel on LEFT side of each scenario tab (cols A-B): toggle Y/N to include/exclude schools")
    print(f"    • Config tab also has master Y/N list as a reference")
    print(f"    • Formulas auto-update instantly when you change Y/N — no re-run needed")
    print(f"    • Re-run this script after editing a school's budget file")
    print(f"    • To add school #{N_SCHOOLS+1}: append to SCHOOLS list → re-run\n")


if __name__ == "__main__":
    main()
